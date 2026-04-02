from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import load_workbook
import random

app = Flask(__name__)

vocab_list = []  # lưu từ vựng


@app.route('/', methods=['GET', 'POST'])
def index():
    global vocab_list

    if request.method == 'POST':
        file = request.files['file']

        if file:
            wb = load_workbook(file)
            ws = wb.active

            vocab_list = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                vocab_list.append({
                    "hanzi": row[2],
                    "pinyin": row[3],
                    "meaning": row[4]
                })

    return render_template('index.html', count=len(vocab_list))


# 🔥 Flashcard
@app.route('/flashcard')
def flashcard():
    if not vocab_list:
        return redirect(url_for('index'))

    word = random.choice(vocab_list)
    return render_template('flashcard.html', word=word)


# 🔥 Quiz: Việt → Trung
@app.route('/quiz', methods=['GET', 'POST'])
def quiz():
    if not vocab_list:
        return redirect(url_for('index'))

    if request.method == 'POST':
        answer = request.form['answer']
        correct = request.form['correct']

        result = (answer.strip() == correct.strip())

        return render_template('quiz.html', result=result)

    word = random.choice(vocab_list)
    return render_template('quiz.html', word=word)


if __name__ == '__main__':
    app.run(debug=True)


<!DOCTYPE html>
<html>
<head>
    <title>Flashcard</title>
    <style>
        body {
            font-family: Arial;
            background: linear-gradient(135deg, #667eea, #764ba2);
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }

        .container {
            text-align: center;
        }

        .card {
            width: 300px;
            height: 200px;
            perspective: 1000px;
            margin-bottom: 20px;
        }

        .inner {
            width: 100%;
            height: 100%;
            position: relative;
            transition: transform 0.6s;
            transform-style: preserve-3d;
            cursor: pointer;
        }

        .card.flip .inner {
            transform: rotateY(180deg);
        }

        .front, .back {
            position: absolute;
            width: 100%;
            height: 100%;
            border-radius: 15px;
            background: white;
            display: flex;
            justify-content: center;
            align-items: center;
            flex-direction: column;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            backface-visibility: hidden;
        }

        .front {
            font-size: 28px;
            font-weight: bold;
        }

        .back {
            transform: rotateY(180deg);
            font-size: 18px;
        }

        .pinyin {
            color: #666;
            margin-top: 10px;
        }

        .btn {
            padding: 10px 20px;
            background: #00c853;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
        }

        .btn:hover {
            background: #00b342;
        }
    </style>
</head>
<body>

<div class="container">

    <div class="card" onclick="flipCard()">
        <div class="inner" id="cardInner">

            <!-- Mặt trước -->
            <div class="front">
                {{ word.hanzi }}
            </div>

            <!-- Mặt sau -->
            <div class="back">
                <div>{{ word.meaning }}</div>
                <div class="pinyin">{{ word.pinyin }}</div>
            </div>

        </div>
    </div>

    <button class="btn" onclick="nextCard()">Next</button>

</div>

<script>
function flipCard() {
    document.querySelector('.card').classList.toggle('flip');
}

function nextCard() {
    window.location.href = "/flashcard";
}
</script>

</body>
</html>


<!DOCTYPE html>
<html>
<head>
    <title>Quiz</title>
    <style>
        body {
            font-family: Arial;
            background: #f5f6fa;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }

        .card {
            background: white;
            padding: 40px;
            border-radius: 15px;
            width: 350px;
            text-align: center;
            box-shadow: 0 5px 20px rgba(0,0,0,0.1);
        }

        h2 {
            margin-bottom: 20px;
        }

        .question {
            font-size: 22px;
            margin-bottom: 20px;
            font-weight: bold;
        }

        input {
            width: 80%;
            padding: 10px;
            font-size: 16px;
            border: 1px solid #ccc;
            border-radius: 8px;
            margin-bottom: 15px;
        }

        button {
            padding: 10px 20px;
            background: #4CAF50;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
        }

        button:hover {
            background: #45a049;
        }

        .correct {
            color: green;
            font-weight: bold;
        }

        .wrong {
            color: red;
            font-weight: bold;
        }

        .next {
            display: inline-block;
            margin-top: 15px;
            text-decoration: none;
            color: white;
            background: #007bff;
            padding: 8px 15px;
            border-radius: 8px;
        }
    </style>
</head>
<body>

<div class="card">
    <h2>Quiz</h2>

    {% if result is defined %}
        {% if result %}
            <p class="correct">✔ Đúng!</p>
        {% else %}
            <p class="wrong">✘ Sai!</p>
            <p>Đáp án: <b>{{ request.form.correct }}</b></p>
        {% endif %}

        <a href="/quiz" class="next">Câu tiếp →</a>

    {% else %}

        <p class="question">{{ word.meaning }}</p>

        <form method="POST">
            <input type="text" name="answer" placeholder="Nhập chữ Hán" autofocus>
            <input type="hidden" name="correct" value="{{ word.hanzi }}">
            <br>
            <button type="submit">Kiểm tra</button>
        </form>

    {% endif %}
</div>

</body>
</html>


<!DOCTYPE html>
<html>
<head>
    <title>Chinese Learning App</title>
    <style>
        body {
            font-family: Arial;
            background: linear-gradient(135deg, #74ebd5, #9face6);
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }

        .card {
            background: white;
            padding: 40px;
            border-radius: 15px;
            width: 350px;
            text-align: center;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }

        h2 {
            margin-bottom: 20px;
        }

        input[type="file"] {
            margin: 15px 0;
        }

        button {
            padding: 10px 20px;
            background: #4CAF50;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
        }

        button:hover {
            background: #45a049;
        }

        .info {
            margin-top: 15px;
            color: #555;
        }

        .links {
            margin-top: 20px;
        }

        .links a {
            display: block;
            margin: 10px 0;
            text-decoration: none;
            background: #007bff;
            color: white;
            padding: 10px;
            border-radius: 8px;
        }

        .links a:hover {
            background: #0056b3;
        }
    </style>
</head>
<body>

<div class="card">
    <h2>📚 Học tiếng Trung</h2>

    <form method="POST" enctype="multipart/form-data">
        <input type="file" name="file">
        <br>
        <button type="submit">Upload từ vựng</button>
    </form>

    <p class="info">Đã load: <b>{{ count }}</b> từ</p>

    {% if count > 0 %}
    <div class="links">
        <a href="/flashcard">🧠 Học Flashcard</a>
        <a href="/quiz">✍️ Làm Quiz</a>
    </div>
    {% endif %}
</div>

</body>
</html>



from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
import random

app = Flask(__name__)

vocab_list = []  # lưu từ vựng


@app.route('/', methods=['GET', 'POST'])
def index():
    global vocab_list

    if request.method == 'POST':
        file = request.files['file']

        if file:
            wb = load_workbook(file)
            ws = wb.active

            vocab_list = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                vocab_list.append({
                    "hanzi": row[0],
                    "pinyin": row[1],
                    "meaning": row[2]
                })

    return render_template('index.html', count=len(vocab_list))


# 🔥 Flashcard
@app.route('/flashcard')
def flashcard():
    if not vocab_list:
        return redirect(url_for('index'))

    word = random.choice(vocab_list)
    return render_template('flashcard.html', word=word)


# 🔥 Quiz: Việt → Trung
@app.route('/quiz', methods=['GET', 'POST'])
def quiz():
    if not vocab_list:
        return redirect(url_for('index'))

    if request.method == 'POST':
        answer = request.form['answer']
        correct = request.form['correct']

        result = (answer.strip() == correct.strip())

        return render_template('quiz.html', result=result)

    word = random.choice(vocab_list)
    return render_template('quiz.html', word=word)


if __name__ == '__main__':
    app.run(debug=True)

<h2>Upload từ vựng</h2>

<form method="POST" enctype="multipart/form-data">
    <input type="file" name="file">
    <button type="submit">Upload</button>
</form>

<p>Đã load: {{ count }} từ</p>

<a href="/flashcard">Học Flashcard</a><br>
<a href="/quiz">Làm Quiz</a>


<h2>Quiz: Nghĩa → Chữ Hán</h2>

{% if result is defined %}
    {% if result %}
        <p style="color: green;">Đúng!</p>
    {% else %}
        <p style="color: red;">Sai!</p>
    {% endif %}
    <a href="/quiz">Câu tiếp</a>
{% else %}

<p>{{ word.meaning }}</p>

<form method="POST">
    <input type="text" name="answer" placeholder="Nhập chữ Hán">
    <input type="hidden" name="correct" value="{{ word.hanzi }}">
    <button type="submit">Kiểm tra</button>
</form>

{% endif %}

<h2>Flashcard</h2>

<h1>{{ word.hanzi }}</h1>

<p>Pinyin: {{ word.pinyin }}</p>
<p>Nghĩa: {{ word.meaning }}</p>

<a href="/flashcard">Next</a>




<input type="file" name="files" webkitdirectory directory multiple>

<form method="POST" enctype="multipart/form-data">
    <input type="file" name="files" webkitdirectory directory multiple>
    <br><br>
    <button type="submit">Upload Folder</button>
</form>
from flask import Flask, render_template, request
from openpyxl import load_workbook
import csv
import io
import re
from collections import defaultdict

app = Flask(__name__)

# 👉 Lấy ngày từ tên file
def get_date_from_filename(filename):
    match = re.search(r'\d{6}', filename)
    if match:
        d = match.group()
        return f"20{d[0:2]}-{d[2:4]}-{d[4:6]}"
    return filename


@app.route('/', methods=['GET', 'POST'])
def index():
    result = None

    if request.method == 'POST':
        files = request.files.getlist('files')

        if not files:
            return render_template('index.html', result=None)

        fail_by_date = defaultdict(int)

        for file in files:
            if not file or file.filename == '':
                continue

            file_date = get_date_from_filename(file.filename)

            try:
                # 👉 Excel
                if file.filename.endswith('.xlsx'):
                    wb = load_workbook(file)
                    ws = wb.active
                    rows = ws.iter_rows(min_row=2, values_only=True)

                # 👉 CSV
                else:
                    content = file.read().decode("utf-8", errors='ignore')
                    stream = io.StringIO(content)
                    reader = csv.reader(stream)
                    next(reader, None)
                    rows = reader

                # 👉 Đếm FAIL trong từng file
                for row in rows:
                    if not row:
                        continue

                    if any('fail' in str(cell).lower() for cell in row):
                        fail_by_date[file_date] += 1

            except Exception as e:
                print("ERROR:", file.filename, e)

        result = dict(fail_by_date)

    return render_template('index.html', result=result)


if __name__ == '__main__':
    app.run(debug=True)

😶
{% if fail_rows and fail_rows|length > 0 %}
<h3>Chi tiết lỗi</h3>
<table>
    <tr>
        <th>Dòng</th>
        <th>Nội dung</th>
    </tr>
    {% for row in fail_rows %}
    <tr>
        <td>{{ row.line }}</td>
        <td>{{ row.data }}</td>
    </tr>
    {% endfor %}
</table>
{% endif %}


from flask import Flask, render_template, request
from openpyxl import load_workbook
import csv
import io
import re

app = Flask(__name__)

# 👉 Lấy ngày từ tên file
def get_date_from_filename(filename):
    match = re.search(r'\d{6}', filename)
    if match:
        d = match.group()
        return f"20{d[0:2]}-{d[2:4]}-{d[4:6]}"
    return filename


@app.route('/', methods=['GET', 'POST'])
def index():
    result = None
    fail_rows = []   # 👉 lưu dòng lỗi

    if request.method == 'POST':
        file = request.files.get('file')

        if not file or file.filename == '':
            return render_template('index.html', result=None, fail_rows=[])

        file_date = get_date_from_filename(file.filename)
        fail_count = 0

        try:
            # 👉 Excel
            if file.filename.endswith('.xlsx'):
                wb = load_workbook(file)
                ws = wb.active
                rows = ws.iter_rows(min_row=2, values_only=True)

            # 👉 CSV
            else:
                content = file.read().decode("utf-8", errors='ignore')
                stream = io.StringIO(content)
                reader = csv.reader(stream)
                next(reader, None)
                rows = reader

            # 👉 Đếm + lưu dòng lỗi
            for idx, row in enumerate(rows, start=2):
                if not row:
                    continue

                if any('fail' in str(cell).lower() for cell in row):
                    fail_count += 1

                    # 👉 lưu dòng lỗi (join lại cho dễ đọc)
                    fail_rows.append({
                        "line": idx,
                        "data": " | ".join(str(x) for x in row)
                    })

            result = {file_date: fail_count}

        except Exception as e:
            result = {"Lỗi": str(e)}

    return render_template('index.html', result=result, fail_rows=fail_rows)


if __name__ == '__main__':
    app.run(debug=True) 


😁
from flask import Flask, render_template, request
from openpyxl import load_workbook
import csv
import io
import re

app = Flask(__name__)

# 👉 Hàm lấy ngày từ tên file
def get_date_from_filename(filename):
    match = re.search(r'\d{6}', filename)
    if match:
        d = match.group()
        return f"20{d[0:2]}-{d[2:4]}-{d[4:6]}"
    return filename  # fallback nếu không match


@app.route('/', methods=['GET', 'POST'])
def index():
    result = None

    if request.method == 'POST':
        file = request.files.get('file')

        if not file or file.filename == '':
            return render_template('index.html', result=None)

        # 👉 lấy ngày từ tên file
        file_date = get_date_from_filename(file.filename)

        fail_count = 0

        try:
            # 👉 Excel
            if file.filename.endswith('.xlsx'):
                wb = load_workbook(file)
                ws = wb.active
                rows = ws.iter_rows(min_row=2, values_only=True)

            # 👉 CSV
            else:
                content = file.read().decode("utf-8", errors='ignore')
                stream = io.StringIO(content)
                reader = csv.reader(stream)
                next(reader, None)
                rows = reader

            # 👉 Đếm FAIL
            for row in rows:
                if not row:
                    continue

                if any('fail' in str(cell).lower() for cell in row):
                    fail_count += 1

            # 👉 Kết quả dạng dict
            result = {file_date: fail_count}

            print("RESULT:", result)

        except Exception as e:
            print("ERROR:", e)
            result = {"Lỗi": str(e)}

    return render_template('index.html', result=result)


if __name__ == '__main__':
    app.run(debug=True)
UnboundLocalError: cannot access local variable 'ws' where it is not associated with a value

Traceback (most recent call last)

from flask import Flask, render_template, request
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import csv
import io

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    result = {}

    if request.method == 'POST':
        file = request.files['file']

        if file:
            fail_count = defaultdict(int)

            # 👉 Excel
            if file.filename.endswith('.xlsx'):
                wb = load_workbook(file)
                ws = wb.active
                rows = ws.iter_rows(min_row=2, values_only=True)

            # 👉 CSV
            else:
                stream = io.StringIO(file.stream.read().decode("utf-8"))
                reader = csv.reader(stream)
                next(reader)
                rows = reader

            for row in rows:
                date = row[0]
                status = str(row[1]).strip().lower()

                if date and status == 'fail':
                    if isinstance(date, datetime):
                        date_str = date.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date)

                    fail_count[date_str] += 1

            result = dict(fail_count)

    return render_template('index.html', result=result)


if __name__ == '__main__':
    app.run(debug=True) 


----------------------
# This is a sample Python script.
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

from flask import Flask, render_template, request
from openpyxl import load_workbook
import os
from collections import defaultdict
app = Flask(__name__)
@app.route('/', methods=['GET','POST'])

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

--------------------------------------
from flask import Flask, render_template, request
from openpyxl import load_workbook
import os
from collections import defaultdict

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/', methods=['GET', 'POST'])
def index():
    result = {}

    if request.method == 'POST':
        file = request.files['file']

        if file:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)

            wb = load_workbook(filepath)
            ws = wb.active

            fail_count = defaultdict(int)

            for row in ws.iter_rows(min_row=2, values_only=True):
                date = row[0]
                status = str(row[1]).strip().lower()

                if date and status == 'fail':
                    date_str = date.strftime('%Y-%m-%d')
                    fail_count[date_str] += 1

            result = dict(fail_count)

    # 👉 ĐẶT Ở ĐÂY
    return render_template('index.html', result=result)
